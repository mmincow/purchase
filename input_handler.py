from datetime import datetime
from pathlib import Path

import openpyxl

from models import PurchaseOrder, OrderItem


def read_from_excel(file_path: str) -> list[PurchaseOrder]:
    """엑셀 파일에서 발주 데이터를 읽어 PurchaseOrder 리스트로 반환"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    orders: dict[str, PurchaseOrder] = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue

        supplier = str(row[0])
        delivery_date = row[1] if isinstance(row[1], datetime) else datetime.strptime(str(row[1]), "%Y-%m-%d")
        project_code = str(row[2] or "")
        account_code = str(row[3] or "")
        item_name = str(row[4])
        specification = str(row[5] or "")
        quantity = int(row[6])
        unit = str(row[7] or "EA")
        unit_price = int(row[8])
        requester = str(row[9] or "")
        department = str(row[10] or "")
        note = str(row[11] or "")

        key = f"{supplier}_{project_code}"

        if key not in orders:
            orders[key] = PurchaseOrder(
                supplier=supplier,
                delivery_date=delivery_date.date() if isinstance(delivery_date, datetime) else delivery_date,
                project_code=project_code,
                account_code=account_code,
                requester=requester,
                department=department,
                note=note,
            )

        orders[key].items.append(
            OrderItem(
                item_name=item_name,
                specification=specification,
                quantity=quantity,
                unit=unit,
                unit_price=unit_price,
            )
        )

    wb.close()
    return list(orders.values())


def create_template(output_path: str = "input/발주요청_템플릿.xlsx"):
    """입력용 엑셀 템플릿 생성"""
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "발주요청"

    headers = [
        "거래처", "납기일", "프로젝트코드", "계정코드",
        "품목명", "규격", "수량", "단위", "단가",
        "요청자", "요청부서", "비고",
    ]
    ws.append(headers)

    ws.append([
        "(주)테스트업체", "2026-07-01", "PJ-001", "41000",
        "볼트 M10x30", "SUS304", 100, "EA", 500,
        "홍길동", "생산팀", "긴급",
    ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 15

    wb.save(output_path)
    wb.close()
    print(f"템플릿 생성 완료: {output_path}")
